import pandas as pd
import time
import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from selenium.webdriver.chrome.options import Options  # Importe as opções do Chrome
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager


def coletarDadosCarbono():
    # Configurando as opções do Chrome para rodar em modo headless
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-cache")
    

    servico = Service(ChromeDriverManager().install())

    navegador = webdriver.Chrome(service=servico, options=chrome_options)
    navegador.get("https://br.investing.com/commodities/carbon-emissions-historical-data")
    navegador.implicitly_wait(60)
    navegador.maximize_window()

    time.sleep(10)
    elementos_tabela = navegador.find_elements(By.XPATH, "//*[@class='w-full text-xs leading-4 overflow-x-auto freeze-column-w-1']")

    if elementos_tabela:
        dados_tabela = []
        for elemento in elementos_tabela:
            linhas = elemento.find_elements(By.TAG_NAME, "tr")
            for linha in linhas:
                colunas = linha.find_elements(By.TAG_NAME, "td")
                dados_linha = [coluna.text for coluna in colunas]
                dados_tabela.append(dados_linha)

        df = pd.DataFrame(dados_tabela, columns=["Data", "Último", "Abertura", "Máxima", "Mínima", "Volume", "Variação"])
        return df
    else:
        print("Tabela não encontrada.")

# Função para formatar a data


def formatar_data(data):
    data_obj = datetime.datetime.strptime(data, "%d.%m.%Y")
    return data_obj.strftime("%d/%m/%Y")


# Chama a função para coletar os dados de Carbono
dados_carbono = coletarDadosCarbono()

# Verifica se os dados foram coletados
if not dados_carbono.empty:
    nome_planilha_saida = "Cotação_Commodities.xlsx"
    abaCarbono = "Carbono"

    try:
        # Carrega o arquivo Excel existente
        book = load_workbook(filename=nome_planilha_saida)
        sheet = book[abaCarbono]
        dados_planilha = pd.DataFrame(sheet.values, columns=["Data", "Último", "Abertura", "Máxima", "Mínima", "Volume", "Variação"])
        dados_novos = dados_carbono[~dados_carbono.isin(
            dados_planilha.to_dict("list"))].dropna()
        linha_inicial = sheet.max_row + 1
    except FileNotFoundError:
        # Se o arquivo não existe, cria um novo com os dados coletados
        book = load_workbook(filename=nome_planilha_saida)
        sheet = book.active
        sheet.title = abaCarbono
        linha_inicial = 1
        dados_novos = dados_carbono

    for row in dados_novos.itertuples(index=False):
        # Formata a data antes de adicionar à planilha
        row = list(row)
        row[0] = formatar_data(row[0])
        sheet.append(row)

    book.save(filename=nome_planilha_saida)

    print(
        f"Dados adicionados com sucesso na aba '{abaCarbono}' da planilha '{nome_planilha_saida}'.")
else:
    print("Não foram encontrados dados para salvar.")
