import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import time
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager


# Função para coletar dados da madeira
def coletarDadosMadeira():
    # Configurações do Chrome
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-cache")

    servico = Service(ChromeDriverManager().install())

    navegador = webdriver.Chrome(service=servico, options=chrome_options)
    navegador.get(
        "https://www.indexmundi.com/pt/pre%C3%A7os-de-mercado/?mercadoria=madeira-dura&meses=180&moeda=brl")
    navegador.implicitly_wait(60)
    navegador.maximize_window()

    time.sleep(10)
    elementos_tabela = navegador.find_elements(By.ID, 'gvPrices')

    if elementos_tabela:
        dados_tabela = []
        for elemento in elementos_tabela:
            linhas = elemento.find_elements(By.TAG_NAME, "tr")
            for linha in linhas:
                colunas = linha.find_elements(By.TAG_NAME, "td")
                dados_linha = [coluna.text for coluna in colunas]
                dados_tabela.append(dados_linha)

        df = pd.DataFrame(dados_tabela, columns=[
                          "Data", "Último", "Taxa Variação"])
        return df
    else:
        print("Tabela não encontrada.")


# Chama a função para coletar os dados da Madeira
dados_madeira = coletarDadosMadeira()

if not dados_madeira.empty:
    nome_planilha_saida = "Cotação_Commodities.xlsx"
    abaMadeira = "Madeira"

    try:
        # Carrega o arquivo Excel existente ou cria um novo
        book = load_workbook(filename=nome_planilha_saida)
        if abaMadeira not in book.sheetnames:
            sheet = book.create_sheet(title=abaMadeira)
        else:
            sheet = book[abaMadeira]

        # Carrega as datas existentes da planilha em um conjunto
        datas_existentes = set([str(row[0]) for row in sheet.iter_rows(
            min_row=2, max_row=sheet.max_row, values_only=True)])

        for row in dados_madeira.itertuples(index=False):
            data = str(row[0])
            if data not in datas_existentes:
                sheet.append(row)
                datas_existentes.add(data)  # Adiciona a nova data ao conjunto

        # Salva as alterações na planilha
        book.save(filename=nome_planilha_saida)

        print(
            f"Dados adicionados com sucesso na aba '{abaMadeira}' da planilha '{nome_planilha_saida}'.")
    except FileNotFoundError:
        print(f"Arquivo '{nome_planilha_saida}' não encontrado.")
else:
    print("Não foram encontrados dados para salvar.")
