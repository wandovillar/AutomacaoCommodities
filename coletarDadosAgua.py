import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl import Workbook
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.select import Select
import re
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager


# Função para formatar o valor removendo /m³ e mantendo apenas números e vírgula
def formatar_valor(valor):
    match = re.search(r'(\d+,\d+)', valor)
    if match:
        return match.group(1)
    return valor.replace(" / m³", "")


def ColetarDadosDaAgua():
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-cache")

    servico = Service(ChromeDriverManager().install())

    navegador = webdriver.Chrome(service=servico, options=chrome_options)
    navegador.get("https://agenciavirtual.sabesp.com.br/web/guest/tarifas")
    navegador.implicitly_wait(60)
    navegador.maximize_window()

    time.sleep(10)

    municipio = navegador.find_element(
        By.ID, "_dxptarifas_WAR_dxptarifas_:tarifasForm:municipios")
    selecao = Select(municipio)
    selecao.select_by_value("100")
    time.sleep(10)

    categoria = navegador.find_element(
        By.ID, "_dxptarifas_WAR_dxptarifas_:tarifasForm:categorias-uso")
    categoriaDeUso = Select(categoria)
    categoriaDeUso.select_by_visible_text("RESIDENCIAL")

    time.sleep(10)

    tarifasAtuais = WebDriverWait(navegador, 10).until(EC.presence_of_element_located(
        (By.ID, "_dxptarifas_WAR_dxptarifas_:tarifasForm:radio-tipos-tarifas:1")))
    navegador.execute_script("arguments[0].click();", tarifasAtuais)

    time.sleep(10)  # Pequeno atraso para dar tempo à página para atualizar

    botao_prosseguir = WebDriverWait(navegador, 10).until(
        EC.presence_of_element_located((By.XPATH, "//*[@class='ui-button-text ui-c']")))
    navegador.execute_script("arguments[0].click();", botao_prosseguir)

    elementos_tabela = WebDriverWait(navegador, 10).until(
        EC.presence_of_all_elements_located((By.XPATH, "(//*[@class='row pl-5'])[2]")))

    if elementos_tabela:
        dados_tabela = []
        for elemento in elementos_tabela:
            linhas = elemento.find_elements(By.TAG_NAME, "tr")
            for linha in linhas:
                colunas = linha.find_elements(By.TAG_NAME, "td")
                dados_linha = [formatar_valor(coluna.text)
                               for coluna in colunas]
                dados_tabela.append(dados_linha)

        df = pd.DataFrame(dados_tabela, columns=[
                          "Categoria", "Faixa de Consumo", "Tarifas de Agua", "Tarifas de Esgoto"])
        return df
    else:
        print("Tabela não encontrada.")


dadosAgua = ColetarDadosDaAgua()

if not dadosAgua.empty:
    nome_planilha_saida = "Cotação_Commodities.xlsx"
    abaAgua = "Agua"

    try:
        book = load_workbook(filename=nome_planilha_saida)
    except FileNotFoundError:
        book = Workbook()

    if abaAgua not in book.sheetnames:
        sheet = book.create_sheet(title=abaAgua)
        sheet.append(["Categoria", "Faixa de Consumo",
                     "Tarifas de Agua", "Tarifas de Esgoto"])
    else:
        sheet = book[abaAgua]

    # Filtrar os dados novos com base na combinação de 'Categoria' e 'Faixa de Consumo'
    dados_planilha = pd.DataFrame(sheet.values)
    dados_planilha.columns = dados_planilha.iloc[0]
    dados_planilha = dados_planilha[1:]
    dados_novos = dadosAgua[~dadosAgua.apply(lambda row: (row['Categoria'], row['Faixa de Consumo']) in dados_planilha[[
                                             'Categoria', 'Faixa de Consumo']].values, axis=1)].dropna()

    for row in dados_novos.itertuples(index=False):
        sheet.append(row)

    book.save(filename=nome_planilha_saida)

    print(
        f"Dados adicionados com sucesso na aba '{abaAgua}' da planilha '{nome_planilha_saida}'.")
else:
    print("Não foram encontrados dados para salvar.")
