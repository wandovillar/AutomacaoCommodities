import datetime
import time
import pandas as pd
import openpyxl 


def adicionarPrecificacaoDolarEuro():
    data_atual = datetime.datetime.now().date() - datetime.timedelta(days=1)
    dia_semana_atual = data_atual.weekday()  # 0: segunda-feira, 6: domingo

    # Verifica se a data é sábado (5) ou domingo (6) e retorna sem adicionar
    if dia_semana_atual == 5 or dia_semana_atual == 6:
        print("Data de Coleta de dados é um sábado ou domingo. Não será coletado valor do Dolar e Euro")
        return

    data_atual_str = data_atual.strftime('%d/%m/%Y')

    arquivo_precificacao = 'PRECIFICACAO ATUAL UCS.xlsx'
    wb_precificacao = openpyxl.load_workbook(arquivo_precificacao)
    ws_precificacao = wb_precificacao['Entrada de Dados']

    linha_inicial = 10
    proxima_linha = linha_inicial

    while ws_precificacao.cell(row=proxima_linha, column=1).value is not None:
        if ws_precificacao.cell(row=proxima_linha, column=1).value == data_atual_str:
            print(f'Valor do Dolar e Euro já adicionada na planilha para a data {data_atual_str}. Volte no próximo dia!')
            return
        proxima_linha += 1

    df_dolar = pd.read_excel('Cotação_Commodities.xlsx', sheet_name='Dolar')
    df_euro = pd.read_excel('Cotação_Commodities.xlsx', sheet_name='Euro')

    valor_dolar = df_dolar.loc[df_dolar['Data'] == data_atual_str, 'Último'].iloc[0]
    valor_euro = df_euro.loc[df_euro['Data'] == data_atual_str, 'Último'].iloc[0]

    ws_precificacao.cell(row=proxima_linha, column=1, value=data_atual_str)
    ws_precificacao.cell(row=proxima_linha, column=2, value=valor_dolar)
    ws_precificacao.cell(row=proxima_linha, column=3, value=valor_euro)

    wb_precificacao.save(arquivo_precificacao)
    print(f'Valores de Dolar e Euro adicionados para a data {data_atual_str} na próxima linha disponível')
    time.sleep(2)
    wb_precificacao.close()
# Chama a função para adicionar os valores do Dolar e Euro na planilha Precificação de UCS
adicionarPrecificacaoDolarEuro()

time.sleep(3)

def adicionarPrecificacaoBoi():
    # Defina o nome do arquivo e da planilha
    arquivo_commodities = 'Cotação_Commodities.xlsx'
    nome_planilha_boi = 'Boi'

    # Carregue a planilha Cotação_Commodities.xlsx
    df = pd.read_excel(arquivo_commodities, sheet_name=nome_planilha_boi)

    # Obtenha a data atual
    data_atual = datetime.datetime.now().date() - datetime.timedelta(days=1)

    # Verifique se a data atual é sábado (dia 5) ou domingo (dia 6)
    if data_atual.weekday() == 5 or data_atual.weekday() == 6:
        print('Data de Coleta de dados é um sábado ou domingo. Não será coletado valor de Boi.')
        return

    data_atual_str = data_atual.strftime('%d/%m/%Y')

    # Abra a planilha Precificação de UCS usando openpyxl
    arquivo_precificacao = 'PRECIFICACAO ATUAL UCS.xlsx'
    wb_precificacao = openpyxl.load_workbook(arquivo_precificacao)
    ws_precificacao = wb_precificacao['Entrada de Dados']

    # Verifique se a data atual já existe na coluna E da planilha
    datas_precificacao = [cell.value for cell in ws_precificacao['E'][10:] if cell.value]
    linha_inicial = 10
    proxima_linha = linha_inicial

    while ws_precificacao.cell(row=proxima_linha, column=5).value is not None:
        if ws_precificacao.cell(row=proxima_linha, column=5).value == data_atual_str:
            print(f'Valor de Boi já adicionado para a data {data_atual_str}. Volte no próximo dia!')
            return
        proxima_linha += 1

    if data_atual_str not in datas_precificacao:
        try:
            valor_boi = df.loc[df['Data'] == data_atual_str, 'Último'].iloc[0]

            ws_precificacao.cell(row=proxima_linha, column=5, value=data_atual_str)
            ws_precificacao.cell(row=proxima_linha, column=6, value=valor_boi)

            wb_precificacao.save(arquivo_precificacao)
            print(f'Valor de Boi adicionado para a data {data_atual_str} na próxima linha disponível')
        except IndexError:
            print(f'Valor de Boi não encontrado para a data {data_atual_str}')
    else:
        print(f'Valor de Boi já foi adicionado. Aguarde o próximo dia!')
        time.sleep(2)

    wb_precificacao.close()  # Feche a planilha após o uso
# Chama a função para adicionar os valores da Boi na planilha Precificação de UCS
adicionarPrecificacaoBoi()

time.sleep(3)

def adicionarPrecificacaoCarbono():
    # Defina o nome do arquivo e da planilha
    arquivo_commodities = 'Cotação_Commodities.xlsx'
    nome_planilha_carbono = 'Carbono'

    # Carregue a planilha Cotação_Commodities.xlsx
    df = pd.read_excel(arquivo_commodities, sheet_name=nome_planilha_carbono)

    # Obtenha a data atual
    data_atual = datetime.datetime.now().date() - datetime.timedelta(days=1)
    
    # Verifique se a data atual é sábado (dia 5) ou domingo (dia 6)
    if data_atual.weekday() == 5 or data_atual.weekday() == 6:
        print('Data de Coleta de dados é um sábado ou domingo. Não será coletado valor de Carbono.')
        return
    
    data_atual_str = data_atual.strftime('%d/%m/%Y')

    # Abra a planilha Precificação de UCS usando openpyxl
    arquivo_precificacao = 'PRECIFICACAO ATUAL UCS.xlsx'
    wb_precificacao = openpyxl.load_workbook(arquivo_precificacao)
    ws_precificacao = wb_precificacao['Entrada de Dados']

    # Verifique se a data atual já existe na coluna W da planilha
    datas_precificacao = [cell.value for cell in ws_precificacao['W'][10:] if cell.value]
    linha_inicial = 10
    proxima_linha = linha_inicial

    while ws_precificacao.cell(row=proxima_linha, column=23).value is not None:
        if ws_precificacao.cell(row=proxima_linha, column=23).value == data_atual_str:
            print(f'Valor de Carbono já adicionado para a data {data_atual_str}. Volte no próximo dia!')
            return
        proxima_linha += 1

    if data_atual_str not in datas_precificacao:
        try:
            valor_carbono = df.loc[df['Data'] == data_atual_str, 'Último'].iloc[0]

            ws_precificacao.cell(row=proxima_linha, column=23, value=data_atual_str)
            ws_precificacao.cell(row=proxima_linha, column=24, value=valor_carbono)

            wb_precificacao.save(arquivo_precificacao)
            print(f'Valor de Carbono adicionado para a data {data_atual_str} na próxima linha disponível')
        except IndexError:
            print(f'Valor de Carbono não encontrado para a data {data_atual_str}')
    else:
        print(f'Valor de Carbono já foi adicionado. Aguarde o próximo dia!')
    time.sleep(2)
    wb_precificacao.close()  # Feche a planilha após o uso
# Chama a função para adicionar os valores da Carbono na planilha Precificação de UCS
adicionarPrecificacaoCarbono()



def adicionarPrecificacaoMilho():
    # Defina o nome do arquivo e da planilha
    arquivo_commodities = 'Cotação_Commodities.xlsx'
    nome_planilha_milho = 'Milho'

    # Carregue a planilha Cotação_Commodities.xlsx
    df = pd.read_excel(arquivo_commodities, sheet_name=nome_planilha_milho)

    # Obtenha a data atual
    data_atual = datetime.datetime.now().date() - datetime.timedelta(days=1)
   
   # Verifique se a data atual é sábado (dia 5) ou domingo (dia 6)
    if data_atual.weekday() == 5 or data_atual.weekday() == 6:
        print('Data de Coleta de dados é um sábado ou domingo. Não será coletado valor do Milho.')
        return
   
    data_atual_str = data_atual.strftime('%d/%m/%Y')

    # Abra a planilha Precificação de UCS usando openpyxl
    arquivo_precificacao = 'PRECIFICACAO ATUAL UCS.xlsx'
    wb_precificacao = openpyxl.load_workbook(arquivo_precificacao)
    ws_precificacao = wb_precificacao['Entrada de Dados']

    # Verifique se a data atual já existe na coluna L da planilha
    datas_precificacao = [cell.value for cell in ws_precificacao['L'][10:] if cell.value]
    linha_inicial = 10
    proxima_linha = linha_inicial

    while ws_precificacao.cell(row=proxima_linha, column=12).value is not None:
        if ws_precificacao.cell(row=proxima_linha, column=12).value == data_atual_str:
            print(f'Valor de Milho já adicionado para a data {data_atual_str}. Volte no próximo dia!')
            return
        proxima_linha += 1

    if data_atual_str not in datas_precificacao:
        try:
            valor_milho = df.loc[df['Data'] == data_atual_str, 'Último'].iloc[0]

            ws_precificacao.cell(row=proxima_linha, column=12, value=data_atual_str)
            ws_precificacao.cell(row=proxima_linha, column=13, value=valor_milho)

            wb_precificacao.save(arquivo_precificacao)
            print(f'Valor de Milho adicionado para a data {data_atual_str} na próxima linha disponível')
        except IndexError:
            print(f'Valor de Milho não encontrado para a data {data_atual_str}')
    else:
        print(f'Valor de Milho já foi adicionado. Aguarde o próximo dia!')
        
    time.sleep(2)
    wb_precificacao.close()  # Feche a planilha após o uso
# Chama a função para adicionar os valores da Milho na planilha Precificação de UCS
adicionarPrecificacaoMilho()



def adicionarPrecificacaoSoja():
    # Defina o nome do arquivo e da planilha
    arquivo_commodities = 'Cotação_Commodities.xlsx'
    nome_planilha_soja = 'Soja'

    # Carregue a planilha Cotação_Commodities.xlsx
    df = pd.read_excel(arquivo_commodities, sheet_name=nome_planilha_soja)

    # Obtenha a data atual
    data_atual = datetime.datetime.now().date() - datetime.timedelta(days=1)
    
    # Verifique se a data atual é sábado (dia 5) ou domingo (dia 6)
    if data_atual.weekday() == 5 or data_atual.weekday() == 6:
        print('Data de Coleta de dados é um sábado ou domingo. Não será coletado valor da Soja.')
        return
    
    data_atual_str = data_atual.strftime('%d/%m/%Y')

    # Abra a planilha Precificação de UCS usando openpyxl
    arquivo_precificacao = 'PRECIFICACAO ATUAL UCS.xlsx'
    wb_precificacao = openpyxl.load_workbook(arquivo_precificacao)
    ws_precificacao = wb_precificacao['Entrada de Dados']

    # Verifique se a data atual já existe na coluna Q da planilha
    datas_precificacao = [cell.value for cell in ws_precificacao['Q'][10:] if cell.value]
    linha_inicial = 10
    proxima_linha = linha_inicial

    while ws_precificacao.cell(row=proxima_linha, column=17).value is not None:
        if ws_precificacao.cell(row=proxima_linha, column=17).value == data_atual_str:
            print(f'Valor de Soja já adicionado para a data {data_atual_str}. Volte no próximo dia!')
            return
        proxima_linha += 1

    if data_atual_str not in datas_precificacao:
        try:
            valor_soja = df.loc[df['Data'] == data_atual_str, 'Último'].iloc[0]

            ws_precificacao.cell(row=proxima_linha, column=17, value=data_atual_str)
            ws_precificacao.cell(row=proxima_linha, column=18, value=valor_soja)

            wb_precificacao.save(arquivo_precificacao)
            print(f'Valor de Soja adicionado para a data {data_atual_str} na próxima linha disponível')
        except IndexError:
            print(f'Valor de Soja não encontrado para a data {data_atual_str}')
    else:
        print(f'Valor de Soja já foi adicionado. Aguarde o próximo dia!')
    time.sleep(2)
    wb_precificacao.close()  # Feche a planilha após o uso
# Chama a função para adicionar os valores da Soja na planilha Precificação de UCS
adicionarPrecificacaoSoja()

time.sleep(3)


def adicionarPrecificacaoMadeira():
    # Defina o nome do arquivo e da planilha de commodities
    arquivo_commodities = 'Cotação_Commodities.xlsx'
    nome_planilha_madeira = 'Madeira'

    # Carregue a planilha Cotação_Commodities.xlsx
    df = pd.read_excel(arquivo_commodities, sheet_name=nome_planilha_madeira)

    # Obtenha a data de ontem
    data_ontem = datetime.datetime.now() - datetime.timedelta(days=1)
    
      # Verifique se a data atual é sábado (dia 5) ou domingo (dia 6)
    if data_ontem.weekday() == 5 or data_ontem.weekday() == 6:
        print('Data de Coleta de dados é um sábado ou domingo. Não será coletado valor da Madeira.')
        return
    
    data_ontem_str = data_ontem.strftime('%d/%m/%Y')

    # Abra a planilha Precificação de UCS usando openpyxl
    arquivo_precificacao = 'PRECIFICACAO ATUAL UCS.xlsx'
    wb_precificacao = openpyxl.load_workbook(arquivo_precificacao)
    ws_precificacao = wb_precificacao['Entrada de Dados']

    # Verifique se a data de ontem já existe na coluna H da planilha de UCS
    datas_precificacao = [cell.value for cell in ws_precificacao['H'][10:] if cell.value]
    linha_inicial = 10
    proxima_linha = linha_inicial

    if data_ontem_str not in datas_precificacao:
        try:
            # Encontre a última atualização da Madeira na planilha Cotação_Commodities
            ultima_atualizacao = df.iloc[-1]
            data_ultima_atualizacao = ultima_atualizacao['Data']
            valor_ultima_atualizacao = ultima_atualizacao['Último']

            # Encontre a próxima linha vazia na coluna H da planilha de UCS
            while ws_precificacao.cell(row=proxima_linha, column=8).value:
                proxima_linha += 1

            # Insira a data de ontem, a data da última atualização e o valor da última atualização
            ws_precificacao.cell(row=proxima_linha, column=8, value=data_ontem_str)
            ws_precificacao.cell(row=proxima_linha, column=9, value=data_ultima_atualizacao)
            ws_precificacao.cell(row=proxima_linha, column=10, value=valor_ultima_atualizacao)

            # Salve a planilha Precificação de UCS
            wb_precificacao.save(arquivo_precificacao)
            print(f'Valores de Madeira adicionados para a data {data_ontem_str} na próxima linha disponível')
        except IndexError:
            print(f'Valores de Madeira não encontrados para a data {data_ontem_str}')
    else:
        print(f'Valores de Madeira para a data {data_ontem_str} já foram adicionados. Volte no próximo dia!')

    wb_precificacao.close()  # Feche a planilha após o uso
    time.sleep(2)
# Chama a função para adicionar os valores da Madeira na planilha Precificação de UCS
adicionarPrecificacaoMadeira()

time.sleep(3)

def adicionarPrecificacaoAgua():
    # Carregar o DataFrame com os dados da planilha de commodities
    arquivo_commodities = 'Cotação_Commodities.xlsx'
    aba_commodities = 'Agua'
    df_commodities = pd.read_excel(
        arquivo_commodities, sheet_name=aba_commodities)

    # Obter o valor da coluna C na linha 8
    valor_agua = df_commodities.at[8, 'Tarifas de Agua']  # Ajuste índice para zero-based

    # Obter a data de ontem
    data_ontem = datetime.datetime.now() - datetime.timedelta(days=1)
    
      # Verifique se a data atual é sábado (dia 5) ou domingo (dia 6)
    if data_ontem.weekday() == 5 or data_ontem.weekday() == 6:
        print('Data de Coleta de dados é um sábado ou domingo. Não será coletado valor da Água.')
        return
    
    data_ontem_str = data_ontem.strftime('%d/%m/%Y')

    # Carregar a planilha Precificação Atual UCS
    arquivo_precificacao = 'PRECIFICACAO ATUAL UCS.xlsx'
    aba_precificacao = 'Entrada de Dados'
    wb_precificacao = openpyxl.load_workbook(arquivo_precificacao)
    ws_precificacao = wb_precificacao[aba_precificacao]

    # Verificar se a data de ontem já existe na coluna 28 (DATA) da planilha
    datas_precificacao = [cell.value for cell in ws_precificacao['AB'][10:] if cell.value]
    linha_inicial = 10
    proxima_linha = linha_inicial

    if data_ontem_str not in datas_precificacao:
        try:
            while ws_precificacao.cell(row=proxima_linha, column=28).value:
                proxima_linha += 1

            # Insira a data de ontem e o valor da coluna C na linha 8
            ws_precificacao.cell(row=proxima_linha, column=28, value=data_ontem_str)
            ws_precificacao.cell(row=proxima_linha, column=29, value=valor_agua)
            wb_precificacao.save(arquivo_precificacao)
            print(f'Valor de água adicionado para a data {data_ontem_str} na próxima linha disponível')
        except IndexError:
            print(f'Valores de água não encontrados para a data {data_ontem_str}')
    else:
        print(f'Valores de água para a data {data_ontem_str} já foram adicionados. Não é necessário duplicar.')

    wb_precificacao.close()
    time.sleep(2)
# Chama a função para adicionar os valores da Água na planilha Precificação de UCS
adicionarPrecificacaoAgua()





